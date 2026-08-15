"""Tests for the five report writers."""

import csv
from datetime import datetime

import pytest
from openpyxl import load_workbook

from app.reports.model import ReportModel, ReportSection, ReportTable
from app.reports.writers import WRITERS, ReportExportError, write_report


PNG_SIGNATURE = b"\x89PNG\r\n\x1a\n"


def make_model(**overrides) -> ReportModel:
    values = {
        "title": "Финансовый отчёт",
        "generated_at": datetime(2026, 8, 15, 19, 30),
        "source": ReportSection(
            title="Источник",
            rows=(("Файл", "transactions.csv"), ("Операций", "500")),
        ),
        "summary": ReportSection(
            title="Показатели",
            rows=(("Выручка", "16 302 389.00"), ("Прибыль", "4 009 296.00")),
        ),
        "insights": ("Прибыль выросла на 11%.", "Средний расход 81 411 в день."),
        "categories": ReportTable(
            title="Категории",
            headers=("Категория", "Операций", "Прибыль"),
            rows=(("Продажи", "120", "10 986 297.00"), ("Аренда", "40", "-2 830 990.00")),
        ),
        "anomalies": ReportTable(
            title="Аномалии",
            headers=("Период", "Показатель", "Тип"),
            rows=(("31.03.2026", "Расходы", "всплеск"),),
        ),
        "periods": ReportTable(
            title="Динамика по периодам",
            headers=("Период", "Выручка"),
            rows=(("01.01.2026", "120 000.00"), ("02.01.2026", "98 000.00")),
        ),
        "chart_png": _real_chart_png(),
        "language": "ru",
        "depth": "detailed",
    }
    values.update(overrides)
    return ReportModel(**values)


def _real_chart_png() -> bytes:
    """Render a genuine PNG so image embedding is exercised for real."""
    from app.analytics.time_series import TimeSeriesPoint, TimeSeriesResult
    from app.reports.chart_image import render_chart_png

    points = [
        TimeSeriesPoint(
            period=datetime(2026, 1, day),
            revenue=100.0 * day,
            expenses=40.0 * day,
            profit=60.0 * day,
            amount=60.0 * day,
        )
        for day in range(1, 6)
    ]
    series = TimeSeriesResult(
        points=points,
        grouping="day",
        week_start="monday",
        visible_series=("revenue", "profit"),
    )
    return render_chart_png(series, None)


ALL_FORMATS = sorted(WRITERS)


@pytest.mark.parametrize("format_key", ALL_FORMATS)
def test_every_writer_produces_a_non_empty_file(tmp_path, format_key):
    target = tmp_path / f"report{WRITERS[format_key].extension}"

    write_report(make_model(), target, format_key)

    assert target.exists()
    assert target.stat().st_size > 0


def test_pdf_has_a_pdf_signature(tmp_path):
    target = tmp_path / "report.pdf"
    write_report(make_model(), target, "pdf")

    assert target.read_bytes()[:5] == b"%PDF-"


def test_png_has_a_png_signature(tmp_path):
    target = tmp_path / "report.png"
    write_report(make_model(), target, "png")

    assert target.read_bytes()[:8] == PNG_SIGNATURE


def test_xlsx_opens_and_carries_every_table(tmp_path):
    target = tmp_path / "report.xlsx"
    write_report(make_model(), target, "xlsx")

    workbook = load_workbook(target)
    names = workbook.sheetnames

    assert "Показатели" in names
    assert "Категории" in names
    assert "Аномалии" in names
    assert "График" in names


def test_xlsx_keeps_cyrillic_values(tmp_path):
    target = tmp_path / "report.xlsx"
    write_report(make_model(), target, "xlsx")

    sheet = load_workbook(target)["Категории"]
    first_value = sheet.cell(row=2, column=1).value

    assert first_value == "Продажи"


def test_csv_parses_and_contains_the_title(tmp_path):
    """The CSV is a stacked document, so it is read row by row, not as a grid."""
    target = tmp_path / "report.csv"
    write_report(make_model(), target, "csv")

    text = target.read_text(encoding="utf-8-sig")
    assert "Финансовый отчёт" in text
    assert "Продажи" in text

    with target.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle, delimiter=";"))

    assert rows[0] == ["Финансовый отчёт"]
    assert ["Категория", "Операций", "Прибыль"] in rows
    assert ["Продажи", "120", "10 986 297.00"] in rows


def test_csv_starts_with_a_utf8_bom_so_excel_opens_cyrillic(tmp_path):
    target = tmp_path / "report.csv"
    write_report(make_model(), target, "csv")

    assert target.read_bytes()[:3] == b"\xef\xbb\xbf"


def test_html_contains_every_section_title(tmp_path):
    target = tmp_path / "report.html"
    write_report(make_model(), target, "html")

    text = target.read_text(encoding="utf-8")
    for title in ("Источник", "Показатели", "Выводы", "Категории", "Аномалии"):
        assert title in text


def test_html_embeds_the_chart_inline(tmp_path):
    target = tmp_path / "report.html"
    write_report(make_model(), target, "html")

    text = target.read_text(encoding="utf-8")
    assert "data:image/png;base64," in text
    assert "http://" not in text
    assert "https://" not in text


def test_html_escapes_dangerous_values(tmp_path):
    model = make_model(
        source=ReportSection(
            title="Источник",
            rows=(("Файл", "<script>alert(1)</script>.csv"),),
        )
    )
    target = tmp_path / "report.html"
    write_report(model, target, "html")

    text = target.read_text(encoding="utf-8")
    assert "<script>alert(1)</script>" not in text
    assert "&lt;script&gt;" in text


# Reports without a chart.


CHARTLESS_FORMATS = [key for key in ALL_FORMATS if not WRITERS[key].needs_chart]


@pytest.mark.parametrize("format_key", CHARTLESS_FORMATS)
def test_writers_cope_with_a_report_that_has_no_chart(tmp_path, format_key):
    model = make_model(chart_png=None, periods=None, anomalies=None)
    target = tmp_path / f"report{WRITERS[format_key].extension}"

    write_report(model, target, format_key)

    assert target.stat().st_size > 0


def test_png_refuses_a_report_without_a_chart(tmp_path):
    model = make_model(chart_png=None)

    with pytest.raises(ReportExportError):
        write_report(model, tmp_path / "report.png", "png")


@pytest.mark.parametrize("format_key", CHARTLESS_FORMATS)
def test_writers_cope_with_an_almost_empty_report(tmp_path, format_key):
    model = make_model(
        insights=(),
        categories=None,
        anomalies=None,
        periods=None,
        chart_png=None,
        summary=ReportSection(title="Показатели", rows=()),
    )
    target = tmp_path / f"report{WRITERS[format_key].extension}"

    write_report(model, target, format_key)

    assert target.stat().st_size > 0


def test_unknown_format_is_rejected(tmp_path):
    with pytest.raises(ReportExportError):
        write_report(make_model(), tmp_path / "report.txt", "txt")


def test_every_format_declares_a_matching_extension():
    for key, report_format in WRITERS.items():
        assert report_format.key == key
        assert report_format.extension.startswith(".")
        assert report_format.file_filter.endswith(f"(*{report_format.extension})")
