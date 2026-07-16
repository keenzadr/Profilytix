"""Fast file preview helpers for Excel and CSV files."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from itertools import islice
from pathlib import Path
from typing import Any


SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
CSV_ENCODINGS = ("utf-8", "utf-8-sig", "cp1251")
CSV_DELIMITERS = (",", ";", "\t", "|")
PREVIEW_ROW_LIMIT = 100
PREVIEW_READ_LIMIT = PREVIEW_ROW_LIMIT + 1
SAMPLE_BYTES = 64 * 1024
HEADER_HINTS = (
    "date",
    "time",
    "amount",
    "sum",
    "total",
    "category",
    "type",
    "status",
    "email",
    "phone",
    "name",
    "id",
    "created",
    "updated",
    "дата",
    "время",
    "сумма",
    "итого",
    "категория",
    "тип",
    "статус",
    "почта",
    "телефон",
    "имя",
    "номер",
)
BOOLEAN_VALUES = {"true", "false", "yes", "no", "да", "нет", "истина", "ложь"}


class FileLoadError(Exception):
    """Raised when a file cannot be previewed as a table."""


@dataclass(frozen=True)
class LoadedFile:
    """Preview rows and metadata for a selected spreadsheet file."""

    path: Path
    column_names: list[str]
    preview_rows: list[list[str]]
    file_size_bytes: int
    total_rows: int | None = None
    encoding: str | None = None
    delimiter: str | None = None
    reader: str = ""

    @property
    def file_name(self) -> str:
        """Return the display name for the loaded file."""
        return self.path.name

    @property
    def column_count(self) -> int:
        """Return the number of detected columns."""
        return len(self.column_names)

    @property
    def preview_row_count(self) -> int:
        """Return the number of rows available in the preview."""
        return len(self.preview_rows)

    @property
    def row_count_text(self) -> str:
        """Return a fast, user-friendly row count label."""
        if self.total_rows is not None and self.total_rows >= self.preview_row_count:
            return str(self.total_rows)
        if self.preview_row_count >= PREVIEW_ROW_LIMIT:
            return f"{PREVIEW_ROW_LIMIT}+ (fast preview)"
        return str(self.preview_row_count)

    @property
    def file_size_text(self) -> str:
        """Return the file size in a readable form."""
        return format_file_size(self.file_size_bytes)


def load_file_preview(file_path: str | Path) -> LoadedFile:
    """Read only the file metadata and first preview rows."""
    path = Path(file_path)
    extension = path.suffix.lower()

    if not path.is_file():
        raise FileLoadError("Selected path is not a readable file.")

    if extension not in SUPPORTED_EXTENSIONS:
        supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
        raise FileLoadError(f"Unsupported file type. Please choose one of: {supported}.")

    if extension == ".csv":
        return _load_csv_preview(path)
    if extension == ".xlsx":
        return _load_xlsx_preview(path)

    return _load_xls_preview(path)


def format_file_size(size_bytes: int) -> str:
    """Format a file size using binary units."""
    size = float(size_bytes)
    for unit in ("B", "KB", "MB", "GB"):
        if size < 1024 or unit == "GB":
            return f"{size:.1f} {unit}" if unit != "B" else f"{int(size)} {unit}"
        size /= 1024
    return f"{size_bytes} B"


def _load_csv_preview(path: Path) -> LoadedFile:
    """Load the first rows of a CSV file with the fastest available reader."""
    file_size = path.stat().st_size
    if file_size == 0:
        return LoadedFile(
            path=path,
            column_names=[],
            preview_rows=[],
            file_size_bytes=file_size,
            total_rows=0,
            encoding=CSV_ENCODINGS[0],
            reader="Polars CSV",
        )

    encoding, sample_text = _detect_csv_encoding(path)
    delimiter = _detect_csv_delimiter(sample_text)

    try:
        return _load_csv_preview_with_polars(path, file_size, encoding, delimiter)
    except FileLoadError:
        return _load_csv_preview_with_pandas(path, file_size, encoding, delimiter)


def _load_csv_preview_with_polars(
    path: Path,
    file_size: int,
    encoding: str,
    delimiter: str,
) -> LoadedFile:
    """Load CSV preview using Polars."""
    try:
        import polars as pl
    except ImportError as error:
        raise FileLoadError(
            "Missing fast CSV dependency: polars. Run `pip install -r requirements.txt`."
        ) from error

    if not hasattr(pl, "read_csv"):
        raise FileLoadError(
            "Polars is installed incorrectly. Reinstall dependencies with "
            "`pip install -r requirements.txt`."
        )

    has_header = True
    try:
        frame = pl.read_csv(
            str(path),
            separator=delimiter,
            encoding=_polars_encoding(encoding),
            n_rows=PREVIEW_READ_LIMIT,
            infer_schema=False,
            infer_schema_length=PREVIEW_READ_LIMIT,
            try_parse_dates=False,
            truncate_ragged_lines=True,
            rechunk=False,
        )
    except Exception as error:
        raise FileLoadError(f"Could not read CSV preview: {error}") from error

    if _headers_look_like_data([str(column) for column in frame.columns]):
        has_header = False
        try:
            frame = pl.read_csv(
                str(path),
                separator=delimiter,
                encoding=_polars_encoding(encoding),
                n_rows=PREVIEW_READ_LIMIT,
                infer_schema=False,
                infer_schema_length=PREVIEW_READ_LIMIT,
                try_parse_dates=False,
                truncate_ragged_lines=True,
                rechunk=False,
                has_header=False,
            )
        except Exception as error:
            raise FileLoadError(f"Could not read CSV preview: {error}") from error

    has_more_rows = frame.height > PREVIEW_ROW_LIMIT
    preview_frame = frame.head(PREVIEW_ROW_LIMIT)
    preview_rows = _rows_from_iterable(preview_frame.iter_rows())
    total_rows = None if has_more_rows else preview_frame.height
    column_names = (
        [str(column) for column in preview_frame.columns]
        if has_header
        else _generic_column_names(preview_frame.width)
    )

    return LoadedFile(
        path=path,
        column_names=column_names,
        preview_rows=preview_rows,
        file_size_bytes=file_size,
        total_rows=total_rows,
        encoding=encoding,
        delimiter=_display_delimiter(delimiter),
        reader="Polars CSV",
    )


def _load_csv_preview_with_pandas(
    path: Path,
    file_size: int,
    encoding: str,
    delimiter: str,
) -> LoadedFile:
    """Fallback CSV preview reader using pandas without loading the whole file."""
    try:
        import pandas as pd
    except ImportError as error:
        raise FileLoadError(
            "Missing CSV dependency: pandas. Run `pip install -r requirements.txt`."
        ) from error

    has_header = True
    try:
        frame = pd.read_csv(
            path,
            sep=delimiter,
            encoding=encoding,
            nrows=PREVIEW_READ_LIMIT,
            dtype=str,
            keep_default_na=False,
            engine="c",
            on_bad_lines="skip",
        )
    except Exception as error:
        raise FileLoadError(f"Could not read CSV preview: {error}") from error

    if _headers_look_like_data([str(column) for column in frame.columns]):
        has_header = False
        try:
            frame = pd.read_csv(
                path,
                sep=delimiter,
                encoding=encoding,
                nrows=PREVIEW_READ_LIMIT,
                dtype=str,
                keep_default_na=False,
                engine="c",
                on_bad_lines="skip",
                header=None,
            )
        except Exception as error:
            raise FileLoadError(f"Could not read CSV preview: {error}") from error

    has_more_rows = len(frame) > PREVIEW_ROW_LIMIT
    preview_frame = frame.head(PREVIEW_ROW_LIMIT)
    total_rows = None if has_more_rows else len(preview_frame)
    column_names = (
        [str(column) for column in preview_frame.columns]
        if has_header
        else _generic_column_names(len(preview_frame.columns))
    )

    return LoadedFile(
        path=path,
        column_names=column_names,
        preview_rows=_rows_from_iterable(preview_frame.itertuples(index=False, name=None)),
        file_size_bytes=file_size,
        total_rows=total_rows,
        encoding=encoding,
        delimiter=_display_delimiter(delimiter),
        reader="pandas CSV fallback",
    )


def _load_xlsx_preview(path: Path) -> LoadedFile:
    """Load the first rows of an XLSX file using openpyxl read-only mode."""
    file_size = path.stat().st_size

    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise FileLoadError(
            "Missing Excel dependency: openpyxl. Run `pip install -r requirements.txt`."
        ) from error

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise FileLoadError(f"Could not open Excel file: {error}") from error

    try:
        if not workbook.worksheets:
            return LoadedFile(
                path=path,
                column_names=[],
                preview_rows=[],
                file_size_bytes=file_size,
                total_rows=0,
                reader="openpyxl read-only",
            )

        sheet = workbook.worksheets[0]
        row_iterator = sheet.iter_rows(values_only=True)
        header = next(row_iterator, None)
        if header is None:
            return LoadedFile(
                path=path,
                column_names=[],
                preview_rows=[],
                file_size_bytes=file_size,
                total_rows=0,
                reader="openpyxl read-only",
            )

        column_count = max(sheet.max_column or 0, len(header))
        header_values = _normalize_row(header, column_count)
        if _headers_look_like_data(header_values):
            column_names = _generic_column_names(column_count)
            preview_rows = [header_values]
            preview_rows.extend(
                _normalize_row(row, column_count)
                for row in islice(row_iterator, PREVIEW_ROW_LIMIT - 1)
            )
            total_rows = sheet.max_row
        else:
            column_names = _column_names_from_header(header, column_count)
            preview_rows = [
                _normalize_row(row, column_count)
                for row in islice(row_iterator, PREVIEW_ROW_LIMIT)
            ]
            total_rows = max((sheet.max_row or 1) - 1, 0)

        return LoadedFile(
            path=path,
            column_names=column_names,
            preview_rows=preview_rows,
            file_size_bytes=file_size,
            total_rows=total_rows,
            reader="openpyxl read-only",
        )
    finally:
        workbook.close()


def _load_xls_preview(path: Path) -> LoadedFile:
    """Load the first rows of an XLS file without reading the whole workbook."""
    file_size = path.stat().st_size

    try:
        import pandas as pd
    except ImportError as error:
        raise FileLoadError(
            "Missing Excel dependency: pandas. Run `pip install -r requirements.txt`."
        ) from error

    has_header = True
    try:
        frame = pd.read_excel(path, sheet_name=0, nrows=PREVIEW_READ_LIMIT)
    except Exception as error:
        raise FileLoadError(f"Could not read Excel preview: {error}") from error

    if _headers_look_like_data([str(column) for column in frame.columns]):
        has_header = False
        try:
            frame = pd.read_excel(path, sheet_name=0, nrows=PREVIEW_READ_LIMIT, header=None)
        except Exception as error:
            raise FileLoadError(f"Could not read Excel preview: {error}") from error

    has_more_rows = len(frame) > PREVIEW_ROW_LIMIT
    preview_frame = frame.head(PREVIEW_ROW_LIMIT)
    total_rows = None if has_more_rows else len(preview_frame)
    column_names = (
        [str(column) for column in preview_frame.columns]
        if has_header
        else _generic_column_names(len(preview_frame.columns))
    )

    return LoadedFile(
        path=path,
        column_names=column_names,
        preview_rows=_rows_from_iterable(preview_frame.itertuples(index=False, name=None)),
        file_size_bytes=file_size,
        total_rows=total_rows,
        reader="pandas Excel preview",
    )


def _detect_csv_encoding(path: Path) -> tuple[str, str]:
    """Detect a supported CSV encoding from a small byte sample."""
    sample_bytes = _read_sample_bytes(path)
    last_error: UnicodeDecodeError | None = None

    for encoding in CSV_ENCODINGS:
        try:
            return encoding, sample_bytes.decode(encoding)
        except UnicodeDecodeError as error:
            last_error = error

    message = "Could not decode CSV sample with utf-8, utf-8-sig, or cp1251."
    if last_error is not None:
        message = f"{message} Last error: {last_error}"
    raise FileLoadError(message)


def _read_sample_bytes(path: Path) -> bytes:
    """Read a small sample from the beginning of the file."""
    with path.open("rb") as file:
        return file.read(SAMPLE_BYTES)


def _detect_csv_delimiter(sample_text: str) -> str:
    """Detect a CSV delimiter using a small text sample."""
    try:
        dialect = csv.Sniffer().sniff(sample_text, delimiters=CSV_DELIMITERS)
        return dialect.delimiter
    except csv.Error:
        return _fallback_delimiter(sample_text)


def _fallback_delimiter(sample_text: str) -> str:
    """Choose the most frequent delimiter from the first non-empty lines."""
    lines = [line for line in sample_text.splitlines()[:20] if line.strip()]
    if not lines:
        return ","

    delimiter_counts = {
        delimiter: sum(line.count(delimiter) for line in lines)
        for delimiter in CSV_DELIMITERS
    }
    delimiter, count = max(delimiter_counts.items(), key=lambda item: item[1])
    return delimiter if count > 0 else ","


def _polars_encoding(encoding: str) -> str:
    """Map common Python encoding names to names accepted by Polars."""
    if encoding in {"utf-8", "utf-8-sig"}:
        return "utf8"
    return encoding


def _display_delimiter(delimiter: str) -> str:
    """Return a readable delimiter label."""
    if delimiter == "\t":
        return "tab"
    return delimiter


def _generic_column_names(column_count: int) -> list[str]:
    """Create generic column names for files without a header row."""
    return [f"Column {index + 1}" for index in range(column_count)]


def _headers_look_like_data(headers: list[str]) -> bool:
    """Return whether detected headers look like the first data row."""
    non_empty_headers = [header.strip() for header in headers if header.strip()]
    if not non_empty_headers:
        return False

    normalized_headers = [_normalize_header(header) for header in non_empty_headers]
    if any(_contains_header_hint(header) for header in normalized_headers):
        return False

    data_like_count = sum(_looks_like_data_value(header) for header in normalized_headers)
    return data_like_count >= max(1, len(non_empty_headers) // 2)


def _normalize_header(value: str) -> str:
    """Normalize text for lightweight header checks."""
    text = value.casefold().replace("ё", "е")
    text = re.sub(r"[_\-/\\|:;.,()\[\]{}]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _contains_header_hint(value: str) -> bool:
    """Return whether a value contains a known header-like word."""
    return any(f" {hint} " in f" {value} " for hint in HEADER_HINTS)


def _looks_like_data_value(value: str) -> bool:
    """Return whether a header candidate resembles a data value."""
    if not value:
        return False
    if value in BOOLEAN_VALUES:
        return True
    if "@" in value or value.startswith(("http ", "https ", "www ")):
        return True
    if re.match(r"^\d{4}\s\d{1,2}\s\d{1,2}", value):
        return True
    if re.match(r"^\d{1,2}\s\d{1,2}\s\d{2,4}", value):
        return True
    if re.fullmatch(r"[-+()]?[\d\s,.]+", value):
        return True
    return False


def _column_names_from_header(header: tuple[Any, ...], column_count: int) -> list[str]:
    """Create display-safe column names from a spreadsheet header row."""
    names: list[str] = []
    for index in range(column_count):
        value = header[index] if index < len(header) else None
        text = _format_cell(value).strip()
        names.append(text or f"Column {index + 1}")
    return names


def _normalize_row(row: tuple[Any, ...], column_count: int) -> list[str]:
    """Return a row padded or truncated to the expected column count."""
    values = list(row[:column_count])
    if len(values) < column_count:
        values.extend([None] * (column_count - len(values)))
    return [_format_cell(value) for value in values]


def _rows_from_iterable(rows: Any) -> list[list[str]]:
    """Convert row tuples from a tabular library to display-safe strings."""
    return [[_format_cell(value) for value in row] for row in rows]


def _format_cell(value: Any) -> str:
    """Convert a cell value to a safe display string."""
    if value is None:
        return ""

    try:
        if value != value:
            return ""
    except TypeError:
        pass

    return str(value)
