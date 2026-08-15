"""Write a report as an Excel workbook.

One sheet per section that exists, so a reader can pivot the numbers without
first untangling a single flat dump. The chart, when present, gets its own
sheet rather than floating over data cells.
"""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.reports.model import ReportModel, ReportSection, ReportTable
from app.reports.strings import label
from app.reports.writers.errors import ReportExportError


TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)
MAX_COLUMN_WIDTH = 60
MIN_COLUMN_WIDTH = 10
INVALID_SHEET_CHARS = set(r"[]:*?/\\")
MAX_SHEET_NAME = 31


def write(model: ReportModel, path: Path) -> None:
    """Write the report as a multi-sheet workbook."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    _write_summary_sheet(workbook, model)
    for table in model.tables():
        _write_table_sheet(workbook, table)
    _write_chart_sheet(workbook, model)

    if not workbook.sheetnames:
        raise ReportExportError("The report is empty, so there is nothing to write.")

    workbook.save(str(path))


def _write_summary_sheet(workbook: Workbook, model: ReportModel) -> None:
    """Write the title, source, metrics, and insights onto one sheet."""
    sheet = workbook.create_sheet(_sheet_name(label(model.language, "section_summary")))

    sheet.append([model.title])
    sheet.cell(row=1, column=1).font = TITLE_FONT
    sheet.append(
        [
            label(model.language, "generated_at"),
            f"{model.generated_at:%d.%m.%Y %H:%M}",
        ]
    )

    _append_section(sheet, model.source)
    _append_section(sheet, model.summary)

    if model.insights:
        sheet.append([])
        sheet.append([label(model.language, "section_insights")])
        sheet.cell(row=sheet.max_row, column=1).font = HEADER_FONT
        for text in model.insights:
            sheet.append([text])
            sheet.cell(row=sheet.max_row, column=1).alignment = Alignment(wrap_text=True)

    _fit_columns(sheet)


def _append_section(sheet: object, section: ReportSection) -> None:
    """Append one label/value block with a bold heading."""
    if section.is_empty:
        return

    sheet.append([])
    sheet.append([section.title])
    sheet.cell(row=sheet.max_row, column=1).font = HEADER_FONT
    for row_label, value in section.rows:
        sheet.append([row_label, value])


def _write_table_sheet(workbook: Workbook, table: ReportTable) -> None:
    """Write one table onto its own sheet with a frozen header."""
    sheet = workbook.create_sheet(_sheet_name(table.title))
    sheet.append(list(table.headers))
    for column_index in range(1, len(table.headers) + 1):
        sheet.cell(row=1, column=column_index).font = HEADER_FONT

    for row in table.rows:
        sheet.append(list(row))

    sheet.freeze_panes = "A2"
    _fit_columns(sheet)


def _write_chart_sheet(workbook: Workbook, model: ReportModel) -> None:
    """Place the chart image on a sheet of its own."""
    if not model.has_chart:
        return

    sheet = workbook.create_sheet(_sheet_name(label(model.language, "section_chart")))
    image = ExcelImage(BytesIO(model.chart_png))
    sheet.add_image(image, "A1")


def _fit_columns(sheet: object) -> None:
    """Size columns to their content, within sane bounds."""
    widths: dict[int, int] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > widths.get(cell.column, 0):
                widths[cell.column] = length

    for column_index, width in widths.items():
        bounded = max(MIN_COLUMN_WIDTH, min(width + 2, MAX_COLUMN_WIDTH))
        sheet.column_dimensions[get_column_letter(column_index)].width = bounded


def _sheet_name(title: str) -> str:
    """Return a title Excel accepts as a sheet name."""
    cleaned = "".join(" " if char in INVALID_SHEET_CHARS else char for char in title)
    cleaned = cleaned.strip() or "Report"
    return cleaned[:MAX_SHEET_NAME]
